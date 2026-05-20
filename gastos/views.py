from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.db.models import Sum, Count
from django.contrib import messages
from turnos.models import Turno
from .models import Gasto
from .forms import GastoForm
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import datetime


PRECIO_POR_CLIENTE = {
    'Cayla': 10000,
    'Bauza': 35000,
    'Tomi': 15000,
}

class GastosView(View):
    def get(self, request):
        hoy = datetime.date.today()
        fecha_str = request.GET.get('fecha')

        try:
            fecha_filtro = datetime.date.fromisoformat(fecha_str) if fecha_str else hoy
        except ValueError:
            fecha_filtro = hoy

        form = GastoForm()

        recaudacion = Turno.objects.filter(
            fecha=fecha_filtro, atendido=True
        ).aggregate(total=Sum('precio'))['total'] or 0

        gastos = Gasto.objects.filter(fecha=fecha_filtro)
        total_gastos = gastos.aggregate(total=Sum('monto'))['total'] or 0
        neto = recaudacion - total_gastos

        return render(request, 'gastos/gastos.html', {
            'form': form,
            'recaudacion_hoy': recaudacion,
            'gastos_hoy': gastos,
            'total_gastos': total_gastos,
            'neto': neto,
            'hoy': hoy,
            'fecha_filtro': fecha_filtro,
        })

    def post(self, request):
        form = GastoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto registrado.')
        else:
            messages.error(request, 'Completá todos los campos correctamente.')
        return redirect('gastos')


class EliminarGastoView(View):
    def post(self, request, pk):
        gasto = get_object_or_404(Gasto, pk=pk)
        gasto.delete()
        messages.success(request, 'Gasto eliminado.')
        return redirect('gastos')


class ExportarGastosExcelView(View):
    def get(self, request):
        hoy = datetime.date.today()
        primer_dia_mes = hoy.replace(day=1)

        if hoy.month == 12:
            ultimo_dia_mes = hoy.replace(month=12, day=31)
        else:
            ultimo_dia_mes = hoy.replace(month=hoy.month + 1, day=1) - datetime.timedelta(days=1)

        wb = openpyxl.Workbook()

        # ─── HOJA 1: GASTOS ───
        ws = wb.active
        ws.title = "Gastos"

        header_fill = PatternFill("solid", fgColor="5E0C0E")
        subheader_fill = PatternFill("solid", fgColor="F5E6E6")
        center = Alignment(horizontal="center")

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Resumen mensual: {hoy.strftime('%B %Y').capitalize()}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].fill = header_fill
        ws["A1"].alignment = center

        row = 2
        total_mes_recaudado = 0
        total_mes_gastos = 0
        num_semana = 1

        fecha_actual = primer_dia_mes
        while fecha_actual <= ultimo_dia_mes:
            inicio_semana = fecha_actual
            dias_hasta_domingo = 6 - fecha_actual.weekday()
            fin_semana = min(fecha_actual + datetime.timedelta(days=dias_hasta_domingo), ultimo_dia_mes)

            ws.merge_cells(f"A{row}:E{row}")
            ws[f"A{row}"] = f"Semana {num_semana}  ({inicio_semana.strftime('%d/%m')} — {fin_semana.strftime('%d/%m')})"
            ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{row}"].fill = PatternFill("solid", fgColor="7a1012")
            ws[f"A{row}"].alignment = center
            row += 1

            headers = ["Fecha", "Descripción", "Monto gasto", "Recaudado del día", "Neto del día"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = subheader_fill
                cell.alignment = center
            row += 1

            total_semana_recaudado = 0
            total_semana_gastos = 0

            dia = inicio_semana
            while dia <= fin_semana:
                gastos_dia = Gasto.objects.filter(fecha=dia)
                recaudado_dia = Turno.objects.filter(
                    fecha=dia, atendido=True
                ).aggregate(total=Sum('precio'))['total'] or 0
                total_gastos_dia = gastos_dia.aggregate(total=Sum('monto'))['total'] or 0
                neto_dia = recaudado_dia - total_gastos_dia

                total_semana_recaudado += recaudado_dia
                total_semana_gastos += total_gastos_dia

                if gastos_dia.exists():
                    first = True
                    for gasto in gastos_dia:
                        ws.cell(row=row, column=1, value=dia.strftime('%d/%m/%Y') if first else "")
                        ws.cell(row=row, column=2, value=gasto.descripcion)
                        ws.cell(row=row, column=3, value=float(gasto.monto))
                        ws.cell(row=row, column=4, value=float(recaudado_dia) if first else "")
                        ws.cell(row=row, column=5, value=float(neto_dia) if first else "")
                        row += 1
                        first = False
                else:
                    ws.cell(row=row, column=1, value=dia.strftime('%d/%m/%Y'))
                    ws.cell(row=row, column=2, value="Sin gastos")
                    ws.cell(row=row, column=3, value=0)
                    ws.cell(row=row, column=4, value=float(recaudado_dia))
                    ws.cell(row=row, column=5, value=float(neto_dia))
                    row += 1

                dia += datetime.timedelta(days=1)

            ws.cell(row=row, column=1, value=f"Total Semana {num_semana}")
            ws.cell(row=row, column=3, value=float(total_semana_gastos))
            ws.cell(row=row, column=4, value=float(total_semana_recaudado))
            ws.cell(row=row, column=5, value=float(total_semana_recaudado - total_semana_gastos))
            for col in range(1, 6):
                ws.cell(row=row, column=col).font = Font(bold=True)
                ws.cell(row=row, column=col).fill = subheader_fill
            row += 2

            total_mes_recaudado += total_semana_recaudado
            total_mes_gastos += total_semana_gastos
            fecha_actual = fin_semana + datetime.timedelta(days=1)
            num_semana += 1

        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "TOTAL DEL MES"
        ws[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws[f"A{row}"].fill = header_fill
        ws[f"A{row}"].alignment = center
        ws.cell(row=row, column=3, value=float(total_mes_gastos)).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=4, value=float(total_mes_recaudado)).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=row, column=5, value=float(total_mes_recaudado - total_mes_gastos)).font = Font(bold=True, color="FFFFFF")
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = header_fill

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 16

        # ─── HOJA 2: COLOCACIONES ───
        ws2 = wb.create_sheet(title="Colocaciones")

        ws2.merge_cells("A1:D1")
        ws2["A1"] = f"Colocaciones del mes: {hoy.strftime('%B %Y').capitalize()}"
        ws2["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws2["A1"].fill = header_fill
        ws2["A1"].alignment = center

        row2 = 2
        num_semana = 1
        total_mes_turnos = 0
        total_mes_colocaciones = 0

        fecha_actual = primer_dia_mes
        while fecha_actual <= ultimo_dia_mes:
            inicio_semana = fecha_actual
            dias_hasta_domingo = 6 - fecha_actual.weekday()
            fin_semana = min(fecha_actual + datetime.timedelta(days=dias_hasta_domingo), ultimo_dia_mes)

            # Encabezado semana
            ws2.merge_cells(f"A{row2}:D{row2}")
            ws2[f"A{row2}"] = f"Semana {num_semana}  ({inicio_semana.strftime('%d/%m')} — {fin_semana.strftime('%d/%m')})"
            ws2[f"A{row2}"].font = Font(bold=True, color="FFFFFF")
            ws2[f"A{row2}"].fill = PatternFill("solid", fgColor="7a1012")
            ws2[f"A{row2}"].alignment = center
            row2 += 1

            # Encabezados columnas
            for col, h in enumerate(["Cliente", "Colocaciones", "Total $", ""], 1):
                cell = ws2.cell(row=row2, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = subheader_fill
                cell.alignment = center
            row2 += 1

            clientes_semana = (
                Turno.objects.filter(fecha__gte=inicio_semana, fecha__lte=fin_semana, atendido=True)
                .values('cliente_de')
                .annotate(turnos=Count('id'))
                .order_by('cliente_de')
            )

            total_semana_turnos = 0
            total_semana_colocaciones = 0

            for c in clientes_semana:
                nombre = c['cliente_de'] or 'Sin cliente'
                turnos = c['turnos']
                total = turnos * PRECIO_POR_CLIENTE.get(c['cliente_de'], 0)
                ws2.cell(row=row2, column=1, value=nombre)
                ws2.cell(row=row2, column=2, value=turnos)
                ws2.cell(row=row2, column=3, value=float(total))
                row2 += 1
                total_semana_turnos += turnos
                total_semana_colocaciones += total

            # Total semana
            ws2.cell(row=row2, column=1, value=f"Total Semana {num_semana}")
            ws2.cell(row=row2, column=2, value=total_semana_turnos)
            ws2.cell(row=row2, column=3, value=float(total_semana_colocaciones))
            for col in range(1, 4):
                ws2.cell(row=row2, column=col).font = Font(bold=True)
                ws2.cell(row=row2, column=col).fill = subheader_fill
            row2 += 2

            total_mes_turnos += total_semana_turnos
            total_mes_colocaciones += total_semana_colocaciones
            fecha_actual = fin_semana + datetime.timedelta(days=1)
            num_semana += 1

        # Total mes
        ws2.merge_cells(f"A{row2}:A{row2}")
        ws2.cell(row=row2, column=1, value="TOTAL DEL MES")
        ws2.cell(row=row2, column=2, value=total_mes_turnos)
        ws2.cell(row=row2, column=3, value=float(total_mes_colocaciones))
        for col in range(1, 4):
            ws2.cell(row=row2, column=col).font = Font(bold=True, color="FFFFFF")
            ws2.cell(row=row2, column=col).fill = header_fill

        ws2.column_dimensions['A'].width = 16
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 16

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="resumen_{hoy.strftime("%B_%Y")}.xlsx"'
        wb.save(response)
        return response